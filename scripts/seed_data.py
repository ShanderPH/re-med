"""
Data ingestion script: Excel -> Supabase
Re-med project - SUS-SP 2025 data
"""
import openpyxl
import requests
import json

SUPABASE_URL = "https://yjsvmafgrflhklwiznao.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Inlqc3ZtYWZncmZsaGtsd2l6bmFvIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzUxMDUwOTgsImV4cCI6MjA5MDY4MTA5OH0.Bhy7W9m-m52aT-ygDF1UGvB6FjJ_Rc6wZmcRW1WrrRU"

HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=minimal"
}

def upsert(table: str, data: list) -> None:
    """Insert data in batches of 500."""
    batch_size = 500
    total = len(data)
    for i in range(0, total, batch_size):
        batch = data[i:i+batch_size]
        resp = requests.post(
            f"{SUPABASE_URL}/rest/v1/{table}",
            headers={**HEADERS, "Prefer": "return=minimal"},
            data=json.dumps(batch)
        )
        if resp.status_code not in (200, 201):
            print(f"  ERROR batch {i//batch_size + 1}: {resp.status_code} {resp.text[:200]}")
        else:
            print(f"  Inserted {min(i+batch_size, total)}/{total}")

def main():
    wb = openpyxl.load_workbook("Base_dados_sus_sp.xlsx")

    # --- Step 1: Insert program ---
    print("Creating program SUS-SP...")
    resp = requests.post(
        f"{SUPABASE_URL}/rest/v1/programs",
        headers={**HEADERS, "Prefer": "return=representation"},
        data=json.dumps({"name": "SUS-SP", "year": 2025, "active": True})
    )
    if resp.status_code not in (200, 201):
        print(f"  ERROR: {resp.status_code} {resp.text}")
        return
    program_id = resp.json()[0]["id"]
    print(f"  Program ID: {program_id}")

    # --- Step 2: Concorrência ---
    print("\nInserting Concorrência...")
    ws = wb["Concorr\u00eancia"]
    concorrencia_data = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if row[0] is None:
            continue
        concorrencia_data.append({
            "program_id": program_id,
            "especialidade": str(row[0]).strip(),
            "vagas": int(row[1]),
            "inscritos": int(row[2]),
            "concorrencia": float(row[3]) if row[3] is not None else 0.0
        })
    upsert("concorrencia", concorrencia_data)
    print(f"  Total: {len(concorrencia_data)} specialties")

    # --- Step 3: Classificação Final ---
    print("\nInserting Classificação Final...")
    ws2 = wb["Classifica\u00e7\u00e3o Final"]
    class_data = []
    for i, row in enumerate(ws2.iter_rows(values_only=True)):
        if i == 0:
            continue
        if row[0] is None:
            continue
        class_data.append({
            "program_id": program_id,
            "especialidade": str(row[0]).strip(),
            "nome": str(row[1]).strip() if row[1] else "",
            "classificacao": int(row[2]),
            "nota_final": float(row[3]) if row[3] is not None else 0.0,
            "tipo_lista": str(row[4]).strip() if row[4] else ""
        })
    upsert("classificacao_final", class_data)
    print(f"  Total: {len(class_data)} records")

    # --- Step 4: Vagas Selecionadas ---
    print("\nInserting Vagas Selecionadas...")
    ws3 = wb["Vagas Selecionadas"]
    vagas_data = []
    for i, row in enumerate(ws3.iter_rows(values_only=True)):
        if i == 0:
            continue
        if row[0] is None:
            continue
        vagas_data.append({
            "program_id": program_id,
            "especialidade": str(row[0]).strip(),
            "instituicao": str(row[1]).strip() if row[1] else "",
            "nome_selecionado": str(row[2]).strip() if row[2] else "",
            "tipo_vaga": str(row[3]).strip() if row[3] else "",
            "numero_chamada": str(row[4]).strip() if row[4] else ""
        })
    upsert("vagas_selecionadas", vagas_data)
    print(f"  Total: {len(vagas_data)} records")

    print("\n✓ Seed complete!")

if __name__ == "__main__":
    main()
